from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter
from datetime import datetime

def generate_emi_chart(loan_amount, annual_interest_rate, tenure_years, start_date, file_path):
    monthly_interest_rate = annual_interest_rate / 12 / 100
    total_months = tenure_years * 12
    emi = round(
        loan_amount * monthly_interest_rate * (1 + monthly_interest_rate) ** total_months /
        ((1 + monthly_interest_rate) ** total_months - 1)
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "EMI Chart"
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # Font styles
    default_font = Font(name="Bookman Old Style", size=12)
    bold_font = Font(name="Bookman Old Style", size=12, bold=True)
    title_font = Font(name="Bookman Old Style", size=14, bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill("solid", fgColor="DDEBF7")

    # Regular ₹ style
    rupee_style = NamedStyle(name="rupee_style")
    rupee_style.number_format = '₹#,##0.00'
    rupee_style.font = default_font
    rupee_style.alignment = right_align

    # Bold ₹ style for totals
    rupee_bold = NamedStyle(name="rupee_bold")
    rupee_bold.number_format = '₹#,##0.00'
    rupee_bold.font = bold_font
    rupee_bold.alignment = right_align

    for style in ["rupee_style", "rupee_bold"]:
        if style not in wb.named_styles:
            wb.add_named_style(eval(style))

    # Header title
    ws.merge_cells("A1:E1")
    ws["A1"] = "EMI Chart"
    ws["A1"].font = title_font
    ws["A1"].alignment = center_align

    # Loan Info
    ws["A2"] = "Amount Borrowed"
    ws["A2"].font = bold_font
    ws["B2"] = f"₹ {loan_amount:,.0f}"
    ws["B2"].font = default_font
    ws["C2"] = "Rate Of Interest"
    ws["C2"].font = bold_font
    ws["D2"] = f"{annual_interest_rate}%"
    ws["D2"].font = default_font
    ws.merge_cells("D2:E2")

    # Table Headers
    headers = ["Month", "Principal Paid", "Interest Charged", "Total Payment", "Balance"]
    start_row = 4
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_num, value=header)
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = header_fill

    # EMI calculation
    balance = loan_amount
    row = start_row + 1
    monthly_data = []

    for month_index in range(total_months):
        current_date = start_date.replace(
            year=start_date.year + (start_date.month + month_index - 1) // 12,
            month=(start_date.month + month_index - 1) % 12 + 1
        )
        interest = round(balance * monthly_interest_rate)
        principal = emi - interest
        balance -= principal

        monthly_data.append({
            "date": current_date,
            "principal": principal,
            "interest": interest,
            "payment": emi,
            "balance": max(balance, 0)
        })

    # Group by Financial Year
    fy_group = {}
    for entry in monthly_data:
        fy_year = entry["date"].year if entry["date"].month >= 4 else entry["date"].year - 1
        fy_group.setdefault(fy_year, []).append(entry)

    # Write Data
    for fy_year, entries in fy_group.items():
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.cell(row=row, column=1, value=f"F.Y. {fy_year}-{str(fy_year + 1)[-2:]}").font = bold_font
        ws.cell(row=row, column=1).alignment = center_align
        row += 1

        total_principal = total_interest = total_payment = 0

        for entry in entries:
            ws.cell(row=row, column=1, value=entry["date"].strftime("%b-%y")).font = default_font
            ws.cell(row=row, column=2, value=entry["principal"]).style = "rupee_style"
            ws.cell(row=row, column=3, value=entry["interest"]).style = "rupee_style"
            ws.cell(row=row, column=4, value=entry["payment"]).style = "rupee_style"
            ws.cell(row=row, column=5, value=entry["balance"]).style = "rupee_style"
            row += 1

            total_principal += entry["principal"]
            total_interest += entry["interest"]
            total_payment += entry["payment"]

        # Totals Row
        ws.cell(row=row, column=1, value="Total").font = bold_font
        ws.cell(row=row, column=1).alignment = center_align
        ws.cell(row=row, column=2, value=total_principal).style = "rupee_bold"
        ws.cell(row=row, column=3, value=total_interest).style = "rupee_bold"
        ws.cell(row=row, column=4, value=total_payment).style = "rupee_bold"
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = thin_border
        row += 2

    # Set Column Widths
    col_widths = [15, 20, 20, 20, 20]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    wb.save(file_path)


from openpyxl import load_workbook
from reportlab.pdfgen.canvas import Canvas
from reportlab.lib.pagesizes import A4
from reportlab.platypus import Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# Register fonts
pdfmetrics.registerFont(TTFont("Bookman", "BOOKOS.TTF"))
pdfmetrics.registerFont(TTFont("Bookman-Bold", "BOOKOSB.TTF"))
import sys
import os

if getattr(sys, 'frozen', False):
    base_path = sys._MEIPASS
else:
    base_path = os.path.abspath(".")

pdfmetrics.registerFont(TTFont("DejaVu", os.path.join(base_path, "DejaVuSans.ttf")))
pdfmetrics.registerFont(TTFont("Bookman", os.path.join(base_path, "BOOKOS.TTF")))
pdfmetrics.registerFont(TTFont("Bookman-Bold", os.path.join(base_path, "BOOKOSB.TTF")))


def generate_pdf_from_excel(excel_path, pdf_path, loan_amount, interest_rate):
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    canvas = Canvas(pdf_path, pagesize=A4)
    width, height = A4
    title_font_size = 16
    header_font_size = 10
    row_font_size = 9

    row = 4
    y_start = height - 2 * cm
    is_first_page = True

    while row <= ws.max_row:
        # Fetch the FY title from current row, but don't use "Month"
        current_cell = ws.cell(row=row, column=1).value

        if not current_cell or current_cell == "Month":
            row += 1
            continue  # Skip invalid rows

        fy_title = current_cell
        row += 1  # Move past FY label row

        # Build table data
        data = [["Month", "Principal Paid", "Interest Charged", "Total Payment", "Balance"]]
        while row <= ws.max_row and ws.cell(row=row, column=1).value != "Total":
            data.append([
                str(ws.cell(row=row, column=1).value),
                f"₹{ws.cell(row=row, column=2).value:,.0f}" if ws.cell(row=row, column=2).value else "",
                f"₹{ws.cell(row=row, column=3).value:,.0f}" if ws.cell(row=row, column=3).value else "",
                f"₹{ws.cell(row=row, column=4).value:,.0f}" if ws.cell(row=row, column=4).value else "",
                f"₹{ws.cell(row=row, column=5).value:,.0f}" if ws.cell(row=row, column=5).value else "",
            ])
            row += 1

        if row <= ws.max_row:
            data.append([
                str(ws.cell(row=row, column=1).value),
                f"₹{ws.cell(row=row, column=2).value:,.0f}" if ws.cell(row=row, column=2).value else "",
                f"₹{ws.cell(row=row, column=3).value:,.0f}" if ws.cell(row=row, column=3).value else "",
                f"₹{ws.cell(row=row, column=4).value:,.0f}" if ws.cell(row=row, column=4).value else "",
                f"₹{ws.cell(row=row, column=5).value:,.0f}" if ws.cell(row=row, column=5).value else "",
            ])
        row += 2

        # Create table
        table = Table(data, colWidths=[3 * cm, 3.5 * cm, 3.5 * cm, 3.5 * cm, 3.5 * cm])
        style = TableStyle([
            ('FONTNAME', (0, 0), (-1, 0), 'Bookman-Bold'),
            ('FONTNAME', (0, 1), (-1, -2), 'Bookman'),
            ('FONTNAME', (0, -1), (-1, -1), 'Bookman-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), header_font_size),
            ('FONTSIZE', (0, 1), (-1, -2), row_font_size),
            ('FONTSIZE', (0, -1), (-1, -1), header_font_size),
            ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
        ])
        for col in range(1, 5):
            style.add('FONTNAME', (col, 1), (col, -2), 'DejaVu')
            style.add('FONTNAME', (col, -1), (col, -1), 'DejaVu')

        table.setStyle(style)
        table.wrapOn(canvas, width, height)
        table_height = table._height
        needed_space = table_height + 2.5 * cm

        # Check for space and handle page break
        if y_start - needed_space < 3 * cm:
            canvas.showPage()
            y_start = height - 2 * cm
            is_first_page = False

        # First page: Title + loan info
        if is_first_page:
            canvas.setFont("Bookman-Bold", title_font_size)
            canvas.drawCentredString(width / 2, y_start, "EMI Chart")
            y_start -= 1.2 * cm
            canvas.setFont("DejaVu", header_font_size)
            canvas.drawString(2 * cm, y_start, f"Amount Borrowed: ₹{loan_amount:,.0f}")
            canvas.drawString(11 * cm, y_start, f"Rate of Interest: {interest_rate:.2f}%")
            y_start -= 1.5 * cm
            is_first_page = False

        # FY title outside table
        # Debug box and log
        print(f"Drawing FY title '{fy_title}' at y = {y_start}")
        canvas.setFont("Bookman-Bold", header_font_size + 2)
        canvas.drawCentredString(width / 2, y_start, str(fy_title))
        y_start -= 1 * cm

        table.drawOn(canvas, 2 * cm, y_start - table_height)
        y_start -= table_height + 2 * cm

    canvas.save()




import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import os

# Global Excel path tracker
last_excel_path = ""

def launch_gui():
    def on_generate_excel():
        global last_excel_path
        try:
            amount = float(entry_amount.get())
            rate = float(entry_rate.get())
            tenure = int(entry_tenure.get())
            start = datetime.strptime(entry_start.get(), "%d-%m-%Y")

            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="Save Excel File"
            )

            if file_path:
                generate_emi_chart(amount, rate, tenure, start, file_path)
                last_excel_path = file_path
                messagebox.showinfo("Success", f"Excel file saved at:\n{file_path}")
                btn_pdf.config(state="normal")
        except ValueError:
            messagebox.showerror("Invalid Input", "Please check inputs and use date format DD-MM-YYYY.")
        except Exception as e:
            messagebox.showerror("Error", f"Something went wrong:\n{e}")

    def on_create_pdf():
        global last_excel_path
        if not last_excel_path or not os.path.exists(last_excel_path):
            messagebox.showerror("Missing File", "Please generate the Excel file first.")
            return
        try:
            amount = float(entry_amount.get())
            rate = float(entry_rate.get())
            pdf_path = os.path.splitext(last_excel_path)[0] + ".pdf"
            generate_pdf_from_excel(last_excel_path, pdf_path, amount, rate)
            messagebox.showinfo("PDF Created", f"PDF saved at:\n{pdf_path}")
        except Exception as e:
            messagebox.showerror("Error", f"PDF creation failed:\n{e}")

    root = tk.Tk()
    root.title("EMI Chart Generator")
    root.geometry("440x360")
    root.configure(bg="#e1ecf4")

    style = ttk.Style()
    style.theme_use("default")
    style.configure("TLabel", background="#e1ecf4", font=("Bookman Old Style", 10))
    style.configure("TButton", font=("Bookman Old Style", 11))
    style.configure("TEntry", font=("Bookman Old Style", 10))

    # Input Fields
    ttk.Label(root, text="Loan Amount (₹):").pack(pady=(10, 2))
    entry_amount = ttk.Entry(root, width=30)
    entry_amount.pack()

    ttk.Label(root, text="Interest Rate (% per annum):").pack(pady=(10, 2))
    entry_rate = ttk.Entry(root, width=30)
    entry_rate.pack()

    ttk.Label(root, text="Tenure (in years):").pack(pady=(10, 2))
    entry_tenure = ttk.Entry(root, width=30)
    entry_tenure.pack()

    ttk.Label(root, text="Start Date (DD-MM-YYYY):").pack(pady=(10, 2))
    entry_start = ttk.Entry(root, width=30)
    entry_start.insert(0, "01-04-2025")
    entry_start.pack()

    # Buttons
    button_frame = tk.Frame(root, bg="#e1ecf4")
    button_frame.pack(pady=25)

    btn_excel = ttk.Button(button_frame, text="Generate Excel File", command=on_generate_excel)
    btn_excel.grid(row=0, column=0, padx=10)

    btn_pdf = ttk.Button(button_frame, text="Create PDF", command=on_create_pdf, state="disabled")
    btn_pdf.grid(row=0, column=1, padx=10)

    root.mainloop()

# Call to start GUI
launch_gui()
